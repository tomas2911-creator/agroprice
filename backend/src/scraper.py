import httpx
import openpyxl
import logging
import io
import re
from datetime import date, timedelta, datetime
from typing import Optional

from src.config import ODEPA_BASE_URL
from src.database import insertar_precios, registrar_importacion, boletin_ya_importado

logger = logging.getLogger("agroprice.scraper")


def generar_url_boletin(fecha: date) -> str:
    """Generar URL del boletín Excel para una fecha dada"""
    anio = fecha.strftime("%Y")
    mes = fecha.strftime("%m")
    fecha_str = fecha.strftime("%Y%m%d")
    return f"{ODEPA_BASE_URL}/{anio}/{mes}/Boletin_Diario_de_Frutas_y_Hortalizas_{fecha_str}.xlsx"


async def descargar_boletin(fecha: date) -> Optional[bytes]:
    """Descargar el archivo Excel del boletín"""
    url = generar_url_boletin(fecha)
    logger.info(f"Descargando boletín: {url}")

    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
        try:
            response = await client.get(url)
            if response.status_code == 200:
                content_type = response.headers.get("content-type", "")
                if "html" in content_type.lower():
                    logger.warning(f"Boletín {fecha}: respuesta HTML en vez de Excel (probablemente no existe)")
                    return None
                logger.info(f"Boletín {fecha}: descargado ({len(response.content)} bytes)")
                return response.content
            elif response.status_code == 404:
                logger.info(f"Boletín {fecha}: no disponible (404)")
                return None
            else:
                logger.warning(f"Boletín {fecha}: HTTP {response.status_code}")
                return None
        except Exception as e:
            logger.error(f"Error descargando boletín {fecha}: {e}")
            return None


def parsear_excel(contenido: bytes, fecha: date) -> list[dict]:
    """
    Parsear el Excel de ODEPA y extraer registros de precios.

    El formato típico del Excel de ODEPA tiene:
    - Múltiples hojas, una por mercado (ej: "Lo Valledor", "Vega Central", etc.)
    - Cada hoja tiene columnas: Producto, Variedad, Unidad, Volumen, Precio Mín, Precio Máx, Precio Prom
    - O puede tener una estructura con encabezados de sección para Frutas y Hortalizas
    """
    registros = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception as e:
        logger.error(f"Error abriendo Excel {fecha}: {e}")
        return []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_name = sheet_name.strip()

        # Ignorar hojas que no son mercados
        if raw_name.lower() in ["resumen", "notas", "info", "hoja1", "sheet1"]:
            continue

        nombre_mercado = _extraer_nombre_mercado(ws, raw_name)
        categoria = _extraer_categoria(ws, raw_name)

        registros_hoja = _parsear_hoja(ws, nombre_mercado, fecha, categoria)
        registros.extend(registros_hoja)
        logger.debug(f"  Hoja '{raw_name}' → mercado='{nombre_mercado}', cat='{categoria}': {len(registros_hoja)} registros")

    # Si no encontramos datos en hojas separadas, intentar formato de hoja única
    if not registros and len(wb.sheetnames) > 0:
        ws = wb[wb.sheetnames[0]]
        registros = _parsear_hoja_unica(ws, fecha)

    wb.close()
    logger.info(f"Boletín {fecha}: {len(registros)} registros extraídos de {len(wb.sheetnames)} hojas")
    return registros


def _limpiar_texto(valor) -> str:
    """Limpiar y normalizar texto de celda"""
    if valor is None:
        return ""
    return str(valor).strip()


def _limpiar_numero(valor, permitir_cero: bool = False) -> Optional[float]:
    """Limpiar y convertir número de celda"""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        if valor == 0:
            return 0.0 if permitir_cero else None
        return float(valor)
    texto = str(valor).strip().replace(".", "").replace(",", ".")
    texto = re.sub(r'[^\d.\-]', '', texto)
    try:
        num = float(texto)
        if num == 0:
            return 0.0 if permitir_cero else None
        return num if num >= 0 else None
    except (ValueError, TypeError):
        return None


def _detectar_columnas(ws) -> dict:
    """
    Detectar las columnas del Excel buscando encabezados conocidos.
    Retorna un dict con mapping de campo -> índice de columna.
    Busca una fila donde cada keyword esté en una CELDA SEPARADA.
    """
    patrones = {
        "producto": ["producto", "especie", "nombre"],
        "variedad": ["variedad", "tipo", "var"],
        "calidad": ["calidad", "calibre", "grado"],
        "unidad": ["unidad", "medida", "envase", "presentación", "presentacion", "comercialización", "comercializacion"],
        "volumen": ["volumen", "vol", "cantidad", "cajas", "unidades"],
        "precio_min": ["mínimo", "minimo", "min", "precio mín", "precio min", "p. min"],
        "precio_max": ["máximo", "maximo", "max", "precio máx", "precio max", "p. max"],
        "precio_promedio": ["promedio", "prom", "precio prom", "p. prom", "ponderado"],
    }

    for row in ws.iter_rows(min_row=1, max_row=15, values_only=False):
        row_mapping = {}
        used_columns = set()

        for cell in row:
            if cell.value is None:
                continue
            texto = str(cell.value).strip().lower()
            col_idx = cell.column - 1  # 0-indexed

            for campo, keywords in patrones.items():
                if campo not in row_mapping and col_idx not in used_columns:
                    for kw in keywords:
                        if kw in texto:
                            row_mapping[campo] = col_idx
                            used_columns.add(col_idx)
                            break

        # Si encontramos al menos producto y algún precio en esta fila, es la fila de headers
        if "producto" in row_mapping and ("precio_promedio" in row_mapping or "precio_min" in row_mapping):
            return row_mapping

    return {}


def _extraer_nombre_mercado(ws, sheet_name: str) -> str:
    """Extraer nombre limpio del mercado desde Row 7 o del nombre de la hoja"""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and row[0]:
            texto = str(row[0]).strip()
            if texto.lower().startswith("mercado:"):
                return texto.split(":", 1)[1].strip()
    # Fallback: limpiar nombre de hoja (quitar prefijo Frutas_ o Hortalizas_)
    clean = re.sub(r'^(Frutas_|Hortalizas_)', '', sheet_name)
    return clean.strip()


def _extraer_categoria(ws, sheet_name: str) -> str:
    """Extraer categoría desde la hoja o nombre de hoja"""
    # Primero intentar desde nombre de hoja
    if sheet_name.lower().startswith("fruta"):
        return "fruta"
    if sheet_name.lower().startswith("hortaliza"):
        return "hortaliza"
    # Fallback: buscar en las primeras filas
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row and row[0]:
            texto = str(row[0]).strip().lower()
            if any(kw in texto for kw in ["fruta", "frutas"]):
                return "fruta"
            if any(kw in texto for kw in ["hortaliza", "hortalizas"]):
                return "hortaliza"
    return "hortaliza"


def _parsear_hoja(ws, mercado: str, fecha: date, categoria: str = None) -> list[dict]:
    """Parsear una hoja que representa un mercado"""
    registros = []
    mapping = _detectar_columnas(ws)

    if "producto" not in mapping:
        return registros

    categoria_actual = categoria or "hortaliza"
    header_row_found = False

    for row in ws.iter_rows(values_only=True):
        valores = list(row)

        if not valores or all(v is None for v in valores):
            continue

        # Detectar sección de categoría
        primer_valor = _limpiar_texto(valores[0]).lower()
        if any(kw in primer_valor for kw in ["fruta", "frutas"]):
            categoria_actual = "fruta"
            continue
        if any(kw in primer_valor for kw in ["hortaliza", "hortalizas", "verdura", "verduras"]):
            categoria_actual = "hortaliza"
            continue

        # Detectar fila de encabezado (saltar)
        if any(kw in primer_valor for kw in ["producto", "especie", "nombre", "mínimo", "máximo"]):
            header_row_found = True
            continue

        # Saltar filas de metadatos
        if any(kw in primer_valor for kw in ["detalle", "día:", "dia:", "precios con", "mercado:", "fuente:"]):
            continue

        # Extraer datos
        producto = _limpiar_texto(valores[mapping.get("producto", 0)])
        if not producto or len(producto) < 2:
            continue

        # Ignorar filas que son subtotales o totales
        if any(kw in producto.lower() for kw in ["total", "subtotal", "suma", "promedio general", "fuente"]):
            continue

        variedad = _limpiar_texto(valores[mapping["variedad"]]) if "variedad" in mapping else None
        calidad = _limpiar_texto(valores[mapping["calidad"]]) if "calidad" in mapping else None
        unidad = _limpiar_texto(valores[mapping["unidad"]]) if "unidad" in mapping else None

        precio_min = _limpiar_numero(valores[mapping["precio_min"]]) if "precio_min" in mapping else None
        precio_max = _limpiar_numero(valores[mapping["precio_max"]]) if "precio_max" in mapping else None
        precio_prom = _limpiar_numero(valores[mapping["precio_promedio"]]) if "precio_promedio" in mapping else None

        volumen = _limpiar_numero(valores[mapping["volumen"]], permitir_cero=True) if "volumen" in mapping else None

        # Solo agregar si hay al menos un precio
        if precio_min is not None or precio_max is not None or precio_prom is not None:
            registros.append({
                "fecha": fecha,
                "mercado": mercado,
                "producto": producto.title(),
                "categoria": categoria_actual,
                "variedad": variedad if variedad else None,
                "calidad": calidad if calidad else None,
                "unidad": unidad if unidad else None,
                "precio_min": precio_min,
                "precio_max": precio_max,
                "precio_promedio": precio_prom,
                "volumen": volumen,
            })

    return registros


def _parsear_hoja_unica(ws, fecha: date) -> list[dict]:
    """Parsear formato de hoja única con columna de mercado"""
    registros = []
    mapping = _detectar_columnas(ws)

    # Buscar columna de mercado
    mercado_col = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            if cell.value and "mercado" in str(cell.value).strip().lower():
                mercado_col = cell.column - 1
                break
        if mercado_col is not None:
            break

    if "producto" not in mapping:
        return registros

    categoria_actual = "hortaliza"

    for row in ws.iter_rows(values_only=True):
        valores = list(row)
        if not valores or all(v is None for v in valores):
            continue

        primer_valor = _limpiar_texto(valores[0]).lower()
        if any(kw in primer_valor for kw in ["fruta", "frutas"]):
            categoria_actual = "fruta"
            continue
        if any(kw in primer_valor for kw in ["hortaliza", "hortalizas"]):
            categoria_actual = "hortaliza"
            continue
        if any(kw in primer_valor for kw in ["producto", "especie", "mercado"]):
            continue

        producto = _limpiar_texto(valores[mapping.get("producto", 0)])
        if not producto or len(producto) < 2:
            continue

        mercado = _limpiar_texto(valores[mercado_col]) if mercado_col is not None else "Desconocido"

        precio_min = _limpiar_numero(valores[mapping["precio_min"]]) if "precio_min" in mapping else None
        precio_max = _limpiar_numero(valores[mapping["precio_max"]]) if "precio_max" in mapping else None
        precio_prom = _limpiar_numero(valores[mapping["precio_promedio"]]) if "precio_promedio" in mapping else None
        volumen = _limpiar_numero(valores[mapping["volumen"]], permitir_cero=True) if "volumen" in mapping else None

        variedad = _limpiar_texto(valores[mapping["variedad"]]) if "variedad" in mapping else None
        calidad = _limpiar_texto(valores[mapping["calidad"]]) if "calidad" in mapping else None
        unidad = _limpiar_texto(valores[mapping["unidad"]]) if "unidad" in mapping else None

        if precio_min is not None or precio_max is not None or precio_prom is not None:
            registros.append({
                "fecha": fecha,
                "mercado": mercado,
                "producto": producto.title(),
                "categoria": categoria_actual,
                "variedad": variedad if variedad else None,
                "calidad": calidad if calidad else None,
                "unidad": unidad if unidad else None,
                "precio_min": precio_min,
                "precio_max": precio_max,
                "precio_promedio": precio_prom,
                "volumen": volumen,
            })

    return registros


async def importar_boletin(fecha: date, forzar: bool = False) -> dict:
    """Descargar e importar un boletín completo"""
    if not forzar and await boletin_ya_importado(fecha):
        return {"fecha": str(fecha), "estado": "ya_importado", "registros": 0}

    contenido = await descargar_boletin(fecha)
    if contenido is None:
        await registrar_importacion(fecha, 0, "no_disponible")
        return {"fecha": str(fecha), "estado": "no_disponible", "registros": 0}

    registros = parsear_excel(contenido, fecha)
    if not registros:
        await registrar_importacion(fecha, 0, "sin_datos", "Excel descargado pero sin datos parseables")
        return {"fecha": str(fecha), "estado": "sin_datos", "registros": 0}

    count = await insertar_precios(registros)
    await registrar_importacion(fecha, count, "ok")

    return {"fecha": str(fecha), "estado": "ok", "registros": count}


async def importar_historico(fecha_inicio: date, fecha_fin: date = None,
                             forzar: bool = False) -> list[dict]:
    """Importar boletines históricos en un rango de fechas"""
    if fecha_fin is None:
        fecha_fin = date.today()

    resultados = []
    fecha_actual = fecha_inicio

    while fecha_actual <= fecha_fin:
        # Solo días hábiles (lunes a viernes)
        if fecha_actual.weekday() < 5:
            resultado = await importar_boletin(fecha_actual, forzar=forzar)
            resultados.append(resultado)
            logger.info(f"Importación {fecha_actual}: {resultado['estado']} ({resultado['registros']} registros)")

        fecha_actual += timedelta(days=1)

    # Resumen
    total_ok = sum(1 for r in resultados if r["estado"] == "ok")
    total_registros = sum(r["registros"] for r in resultados)
    total_no_disp = sum(1 for r in resultados if r["estado"] == "no_disponible")
    total_ya_imp = sum(1 for r in resultados if r["estado"] == "ya_importado")

    logger.info(f"Importación histórica completada: {total_ok} boletines, "
                f"{total_registros} registros, {total_no_disp} no disponibles, "
                f"{total_ya_imp} ya importados")

    return resultados


async def importar_hoy() -> dict:
    """Importar el boletín de hoy"""
    return await importar_boletin(date.today())
